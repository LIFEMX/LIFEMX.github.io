from flask import Flask, render_template, request, redirect
import sqlite3
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

# Crear la base de datos si no existe
def init_db():
    conn = sqlite3.connect('base_de_datos.db')
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS respuestas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            correo TEXT NOT NULL,
            edad INTEGER NOT NULL,
            respuesta TEXT NOT NULL
        )
    ''')
    conn.commit()
    conn.close()

# Crear el archivo Excel si no existe
def init_excel():
    if not os.path.exists('archivo_excel.xlsx'):
        wb = Workbook()
        ws = wb.active
        ws.title = "Respuestas"
        ws.append(['Nombre', 'Correo', 'Edad', 'Respuesta'])
        wb.save('archivo_excel.xlsx')

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        nombre = request.form['nombre']
        correo = request.form['correo']
        edad = request.form['edad']
        respuesta = request.form['respuesta']

        # Guardar en la base de datos
        conn = sqlite3.connect('base_de_datos.db')
        cursor = conn.cursor()
        cursor.execute('INSERT INTO respuestas (nombre, correo, edad, respuesta) VALUES (?, ?, ?, ?)',
                       (nombre, correo, edad, respuesta))
        conn.commit()
        conn.close()

        # Guardar en Excel
        wb = load_workbook('archivo_excel.xlsx')
        ws = wb.active
        ws.append([nombre, correo, edad, respuesta])
        wb.save('archivo_excel.xlsx')

        return redirect('/')
    
    return render_template('form.html')

if __name__ == '__main__':
    init_db()
    init_excel()
    app.run(debug=True)
